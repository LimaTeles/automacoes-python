{
 "cells": [
  {
   "cell_type": "code",
   "execution_count": null,
   "id": "af3ad01c",
   "metadata": {},
   "outputs": [],
   "source": [
    "from openpyxl import load_workbook\n",
    "from openpyxl.styles import PatternFill\n",
    "import os\n",
    "\n",
    "caminho_r1 = '<CAMINHO_DO-ARQUIVO>'\n",
    "caminho_r2 = '<CAMINHO_DO-ARQUIVO>'\n",
    "caminho_r3 = '<CAMINHO_DO-ARQUIVO>'\n",
    "caminho_r4 = '<CAMINHO_DO-ARQUIVO>'\n",
    "caminho_r5 = '<CAMINHO_DO-ARQUIVO>'\n",
    "\n",
    "caminho_meta = '<CAMINHO_DO-ARQUIVO>'\n",
    "caminho_controle = '<CAMINHO_DO-ARQUIVO>'\n",
    "\n",
    "campo_r1 = load_workbook(caminho_r1)\n",
    "campo_r2 = load_workbook(caminho_r2)\n",
    "campo_r3 = load_workbook(caminho_r3)\n",
    "campo_r4 = load_workbook(caminho_r4)\n",
    "campo_r5 = load_workbook(caminho_r5)\n",
    "\n",
    "meta = load_workbook(caminho_meta)\n",
    "controle = load_workbook(caminho_controle)\n",
    "\n",
    "class Atualizacao: \n",
    "    \n",
    "    def __init__(self, sheet_municipio, regiao, nome_municipio):\n",
    "        \n",
    "        self.sheet_municipio = sheet_municipio\n",
    "        self.regiao = regiao\n",
    "        self.nome_municipio = nome_municipio\n",
    "        \n",
    "        self.sheet_meta = meta['SHEET_DA-PLANILHA']\n",
    "        self.sheet_resumo = meta['SHEET_DA-PLANILHA']\n",
    "        self.sheet_micro = controle['SHEET_DA-PLANILHA']\n",
    "        self.sheet_controle = controle['SHEET_DA-PLANILHA']\n",
    "        self.sheet_monitoramento = controle['SHEET_DA-PLANILHA']\n",
    "        \n",
    "    def atualizar_meta(self, coluna_meta):\n",
    "        \n",
    "        dicionario_meta = {}\n",
    "        \n",
    "        for linha in range(2, self.sheet_municipio.max_row + 1):\n",
    "            if all(self.sheet_municipio[coluna + str(linha)].value is not None for coluna in ['L', 'M', 'N', 'O', 'P']):\n",
    "                \n",
    "                data = self.sheet_municipio['L' + str(linha)].value\n",
    "                equipe = self.sheet_municipio['M' + str(linha)].value\n",
    "                \n",
    "                chave_dicionario_meta = (data, equipe)\n",
    "                \n",
    "                if chave_dicionario_meta in dicionario_meta:\n",
    "                    dicionario_meta[chave_dicionario_meta] += 1\n",
    "                else:\n",
    "                    dicionario_meta[chave_dicionario_meta] = 1\n",
    "        \n",
    "        resultados = []\n",
    "        \n",
    "        for (data, equipe), contagem_coletas in dicionario_meta.items():\n",
    "            resultados.append((data, contagem_coletas, equipe))\n",
    "        \n",
    "        proxima_linha = 1\n",
    "        \n",
    "        while self.sheet_meta.cell(row = proxima_linha, column = coluna_meta).value is not None:\n",
    "            proxima_linha += 1   \n",
    "            \n",
    "        for linha, resultado in enumerate(resultados, start = proxima_linha):\n",
    "            for coluna, valor in enumerate(resultado, start = coluna_meta):\n",
    "                self.sheet_meta.cell(row = linha, column = coluna, value = valor)\n",
    "    \n",
    "    def atualizar_controle(self):\n",
    "        \n",
    "        dicionario_controle = {}\n",
    "        \n",
    "        for linha in range(2, self.sheet_municipio.max_row + 1):\n",
    "            if all(self.sheet_municipio[coluna + str(linha)].value is not None for coluna in ['L', 'M', 'N', 'O', 'P']):\n",
    "                \n",
    "                data = self.sheet_municipio['L' + str(linha)].value\n",
    "                equipe = self.sheet_municipio['M' + str(linha)].value\n",
    "                comunidades = self.sheet_municipio['I' + str(linha)].value\n",
    "                \n",
    "                chave_dicionario_controle = (data, equipe, comunidades)\n",
    "                \n",
    "                if chave_dicionario_controle in dicionario_controle:\n",
    "                    dicionario_controle[chave_dicionario_controle] += 1\n",
    "                else:\n",
    "                    dicionario_controle[chave_dicionario_controle] = 1\n",
    "   \n",
    "        resultados = []\n",
    "    \n",
    "        for (data, equipe, comunidades), contagem_coletas in dicionario_controle.items():\n",
    "            resultados.append((data, self.regiao, equipe, self.nome_municipio, comunidades, contagem_coletas))\n",
    "        \n",
    "        proxima_linha = 1\n",
    "        \n",
    "        while self.sheet_controle.cell(row = proxima_linha, column = 2).value is not None:\n",
    "            proxima_linha += 1\n",
    "            \n",
    "        for linha, resultado in enumerate(resultados, start = proxima_linha):\n",
    "            for coluna, valor in enumerate(resultado, start = 1):\n",
    "                self.sheet_controle.cell(row = linha, column = coluna, value = valor)  \n",
    "                \n",
    "    def atualizar_monitoramento(self):\n",
    "        \n",
    "        dicionario_completas = {}\n",
    "        dicionario_incompletas = {}\n",
    "\n",
    "        for linha in range(2, self.sheet_municipio.max_row + 1):\n",
    "            if all(self.sheet_municipio[coluna + str(linha)].value is not None for coluna in ['L', 'M', 'N', 'O', 'P']):\n",
    "\n",
    "                data = self.sheet_municipio['L' + str(linha)].value\n",
    "                equipe = self.sheet_municipio['M' + str(linha)].value\n",
    "                \n",
    "                chave_dicionario_completas = (data, equipe)\n",
    "\n",
    "                if chave_dicionario_completas in dicionario_completas:\n",
    "                    dicionario_completas[chave_dicionario_completas] += 1\n",
    "                else:\n",
    "                    dicionario_completas[chave_dicionario_completas] = 1\n",
    "\n",
    "            if any(self.sheet_municipio[coluna + str(linha)].value is not None for coluna in ['L', 'M', 'N', 'O', 'P']) and any(self.sheet_municipio[coluna + str(linha)].value is None for coluna in ['L', 'M', 'N', 'O', 'P']):\n",
    "\n",
    "                data = self.sheet_municipio['L' + str(linha)].value\n",
    "                equipe = self.sheet_municipio['M' + str(linha)].value\n",
    "                \n",
    "                chave_dicionario_incompletas = (data, equipe)\n",
    "\n",
    "                if chave_dicionario_incompletas in dicionario_incompletas:\n",
    "                    dicionario_incompletas[chave_dicionario_incompletas] += 1\n",
    "                else:\n",
    "                    dicionario_incompletas[chave_dicionario_incompletas] = 1\n",
    "          \n",
    "        resultados = []\n",
    "        \n",
    "        for (data, equipe), contagem_completas in dicionario_completas.items():\n",
    "            resultados.append((data, self.regiao, self.nome_municipio, equipe, contagem_completas, 'Sim'))\n",
    "\n",
    "        for (data, equipe), contagem_incompletas in dicionario_incompletas.items():\n",
    "            resultados.append((data, self.regiao, self.nome_municipio, equipe, contagem_incompletas, 'Não')) \n",
    "\n",
    "        proxima_linha = 1\n",
    "        \n",
    "        while self.sheet_monitoramento.cell(row = proxima_linha, column = 2).value is not None:\n",
    "            proxima_linha += 1\n",
    "            \n",
    "        for linha, resultado in enumerate(resultados, start = proxima_linha):\n",
    "            for coluna, valor in enumerate(resultado, start = 1):\n",
    "                self.sheet_monitoramento.cell(row = linha, column = coluna, value = valor)\n",
    "\n",
    "    def coloracao_incompletas(self):\n",
    "        \n",
    "        for linha in range(2, self.sheet_municipio.max_row + 1):\n",
    "            if any(self.sheet_municipio[coluna + str(linha)].value is not None for coluna in ['L', 'M', 'N', 'O', 'P']) and any(self.sheet_municipio[coluna + str(linha)].value is None for coluna in ['L', 'M', 'N', 'O', 'P']):\n",
    "               \n",
    "                red_fill = PatternFill(start_color = 'FF6666', fill_type = 'solid')\n",
    "                \n",
    "                for coluna in range(12, 17):\n",
    "                    \n",
    "                    self.sheet_municipio.cell(row = linha, column = coluna).fill = red_fill \n",
    "                    \n",
    "    def situacoes(self, linha_resumo):\n",
    "        \n",
    "        endereco_nao_localizado = 0\n",
    "        endereco_outro_municipio = 0\n",
    "        contato_sem_sucesso = 0\n",
    "        \n",
    "        for linha in range(2, self.sheet_municipio.max_row + 1):\n",
    "            \n",
    "            situacao = self.sheet_municipio['S' + str(linha)].value\n",
    "            \n",
    "            if situacao == 'Endereço não localizado':\n",
    "                endereco_nao_localizado +=1\n",
    "            elif situacao == 'Endereço em outro município':\n",
    "                endereco_outro_municipio += 1\n",
    "            elif situacao == 'Contato sem sucesso':\n",
    "                contato_sem_sucesso += 1\n",
    "        \n",
    "        resultados = [endereco_outro_municipio, endereco_nao_localizado, contato_sem_sucesso]\n",
    "            \n",
    "        for coluna, valor in zip('FGH', resultados):\n",
    "            self.sheet_resumo[coluna + str(linha_resumo)] = valor\n",
    "    \n",
    "    def coloracao_monitoramento(self):\n",
    "        \n",
    "        for linha in range(2, self.sheet_monitoramento.max_row + 1):\n",
    "            if self.sheet_monitoramento['F' + str(linha)].value == \"Sim\":\n",
    "\n",
    "                green_fill = PatternFill(start_color = '66FF66', fill_type = 'solid')\n",
    "\n",
    "                for coluna in range(1, 7):\n",
    "\n",
    "                    self.sheet_monitoramento.cell(row = linha, column = coluna).fill = green_fill\n",
    "\n",
    "            if self.sheet_monitoramento['F' + str(linha)].value == \"Não\":\n",
    "\n",
    "                red_fill = PatternFill(start_color = 'FF6666', fill_type = 'solid')\n",
    "\n",
    "                for coluna in range(1, 7):\n",
    "\n",
    "                    self.sheet_monitoramento.cell(row = linha, column = coluna).fill = red_fill\n",
    "     \n",
    "    \n",
    "lista_sheets = [campo_r1['SHEET_DA-PLANILHA'], campo_r2['SHEET_DA-PLANILHA'], campo_r2['SHEET_DA-PLANILHA'], campo_r2['SHEET_DA-PLANILHA'], campo_r2['SHEET_DA-PLANILHA'], campo_r2['SHEET_DA-PLANILHA']]\n",
    "lista_regiao = ['R1', 'R2', 'R2', 'R2', 'R2', 'R2']\n",
    "lista_municipio = ['BRUMADINHO', 'IGARAPÉ', 'BETIM', 'JUATUBA', 'MÁRIO CAMPOS', 'SÃO JOAQUIM DE BICAS']\n",
    "lista_col_meta = [1,4,7,10,13,16]\n",
    "lista_linha_resumo = [2,3,4,5,6,7]\n",
    "\n",
    "for sheet, regiao, municipio, col_meta, linha_resumo in zip(lista_sheets, lista_regiao, lista_municipio, lista_col_meta, lista_linha_resumo):\n",
    "    \n",
    "    atualizacao = Atualizacao(sheet, regiao, municipio)\n",
    "    \n",
    "    atualizacao.atualizar_meta(col_meta)\n",
    "    atualizacao.atualizar_controle()\n",
    "    atualizacao.atualizar_monitoramento()\n",
    "    atualizacao.coloracao_incompletas()\n",
    "    atualizacao.situacoes(linha_resumo)\n",
    "    \n",
    "coloracao_monitoramento = Atualizacao(None, None, None) \n",
    "coloracao_monitoramento.coloracao_monitoramento()\n",
    "\n",
    "campo_r1.save(caminho_r1)\n",
    "campo_r2.save(caminho_r2)\n",
    "campo_r3.save(caminho_r3)\n",
    "campo_r4.save(caminho_r4)\n",
    "campo_r5.save(caminho_r5)            \n",
    "\n",
    "controle.save(caminho_controle)\n",
    "meta.save(caminho_meta)\n",
    "\n",
    "os.startfile(caminho_controle)\n",
    "os.startfile(caminho_meta)\n",
    "\n",
    "print(\"Atualização Concluída\")\n"
   ]
  }
 ],
 "metadata": {
  "kernelspec": {
   "display_name": "Python 3 (ipykernel)",
   "language": "python",
   "name": "python3"
  },
  "language_info": {
   "codemirror_mode": {
    "name": "ipython",
    "version": 3
   },
   "file_extension": ".py",
   "mimetype": "text/x-python",
   "name": "python",
   "nbconvert_exporter": "python",
   "pygments_lexer": "ipython3",
   "version": "3.10.9"
  }
 },
 "nbformat": 4,
 "nbformat_minor": 5
}